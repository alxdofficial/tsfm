"""Generate baseline comparison Excel spreadsheet."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# =========================================================================
# Sheet 1: Baseline Comparison Matrix
# =========================================================================
ws1 = wb.active
ws1.title = "Comparison Matrix"

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
row_header_font = Font(bold=True, size=10)
row_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
our_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
result_font = Font(size=10)
note_font = Font(size=9, italic=True, color="666666")
pending_font = Font(size=10, italic=True, color="CC7722")
incompatible_font = Font(size=10, italic=True, color="CC0000")
nocode_font = Font(size=10, italic=True, color="999999")
retrain_font = Font(size=10, italic=True, color="5B7DAF")
ckpt_font = Font(size=10, italic=True, color="7B5BA6")
impractical_font = Font(size=10, italic=True, color="AAAAAA")
train_font = Font(size=10, italic=True, color="888888")

# Fill colors for status cells
nocode_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
retrain_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
ckpt_fill = PatternFill(start_color="EDE8F5", end_color="EDE8F5", fill_type="solid")
impractical_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Shorthand for empty cells — reproducibility status per baseline
NO_CODE = ("no code, no ckpts", nocode_font, nocode_fill)
LANHAR_RETRAIN = ("has code, no ckpts\nneed retrain+eval", retrain_font, retrain_fill)
CROSSHAR_RETRAIN = ("has code, no ckpts\nneed retrain+eval", retrain_font, retrain_fill)
LIMU_RETRAIN = ("has code, no ckpts for this ds\nneed retrain (supervised only)", retrain_font, retrain_fill)
LIMU_HAS_CKPT = ("has code + ckpts\nneed eval (supervised only)", ckpt_font, ckpt_fill)
IMU2CLIP_NA = ("code archived, no ckpts\nEgo4D only, impractical", impractical_font, impractical_fill)
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")

# Column headers
headers = ["Dataset", "Ours", "NLS-HAR\n(AAAI 2025)", "GOAT\n(IMWUT 2024)",
           "LanHAR\n(IMWUT 2025)", "CrossHAR\n(IMWUT 2024)", "LIMU-BERT\n(SenSys 2021)",
           "IMU2CLIP\n(EMNLP 2023)"]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = thin_border

ws1.row_dimensions[1].height = 36

# Data rows: (dataset, ours, nls_har, goat, lanhar, crosshar, limu_bert, imu2clip)
# Each cell is (text, font_override_or_None)
rows = [
    # (dataset, ours, nls_har, goat, lanhar, crosshar, limu_bert, imu2clip)
    # Each cell: (text, font) for results, or (text, font, fill) for status cells, or None should not appear

    ("MotionSense",
     ("53.72% F1\nclosed-set, zero-shot", result_font),
     ("38.97% F1\nclosed-set, zero-shot", result_font),
     NO_CODE,
     ("76.0% F1\n4-act, single-src (UCI)", result_font),
     ("78.26% acc\n4-act still=sit+stand, multi-src (3ds)", result_font),
     ("89.9% F1\nsupervised, 1% labels", result_font),
     IMU2CLIP_NA),

    ("MobiAct",
     ("19.83% F1\nclosed-set, zero-shot", result_font),
     ("16.93% F1\nclosed-set, zero-shot", result_font),
     NO_CODE,
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("RealWorld",
     ("36.89% F1\nclosed-set, zero-shot", result_font),
     NO_CODE,
     ("78.49% F1\nGOAT-CLIP, few-shot (~7% target)", result_font),
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("Shoaib",
     ("pending\n(in benchmark script)", pending_font),
     NO_CODE,
     NO_CODE,
     ("71.2% F1\n4-act, single-src (UCI)", result_font),
     ("73.67% acc\n4-act still=sit+stand, multi-src (3ds)", result_font),
     ("89.9% F1\nsupervised, 1% labels", result_font),
     IMU2CLIP_NA),

    ("Realdisp",
     ("pending\n(in benchmark script)", pending_font),
     NO_CODE,
     ("81.39% F1\nGOAT-CLIP, few-shot (~7% target)", result_font),
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("Opportunity",
     ("incompatible labels\nwe: 4 locomotion, they: 17 gestures", incompatible_font),
     NO_CODE,
     ("54.35% F1\nGOAT-BERT, few-shot (~7% target)", result_font),
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("Daphnet FoG",
     ("incompatible labels\nwe: 2 classes, they: 3 incl 'Null'", incompatible_font),
     NO_CODE,
     ("64.14% F1\nGOAT-BERT, few-shot (~7% target)", result_font),
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("HHAR",
     ("in training set", train_font),
     ("31.05% F1\nzero-shot", result_font),
     NO_CODE,
     ("80.4% F1\n4-act, single-src (UCI)", result_font),
     ("76.19% acc\n4-act, multi-src (3ds)", result_font),
     ("96.2% F1\nsupervised", result_font),
     IMU2CLIP_NA),

    ("UCI-HAR",
     ("in training set", train_font),
     NO_CODE,
     NO_CODE,
     ("80.8% F1\n4-act, single-src (HHAR)", result_font),
     ("88.68% acc\n4-act, multi-src (3ds)", result_font),
     ("92.3% F1\nsupervised", result_font),
     IMU2CLIP_NA),

    ("MHEALTH",
     ("in training set", train_font),
     ("11.15% F1\nzero-shot", result_font),
     NO_CODE,
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),

    ("PAMAP2",
     ("in training set", train_font),
     ("10.88% F1\nzero-shot", result_font),
     ("77.13% F1\nGOAT-CLIP, few-shot", result_font),
     LANHAR_RETRAIN,
     CROSSHAR_RETRAIN,
     LIMU_RETRAIN,
     IMU2CLIP_NA),
]

for row_idx, row_data in enumerate(rows, 2):
    ds_name = row_data[0]

    # Dataset name cell
    cell = ws1.cell(row=row_idx, column=1, value=ds_name)
    cell.font = row_header_font
    cell.fill = row_header_fill
    cell.border = thin_border
    cell.alignment = Alignment(vertical="top")

    # Data cells
    for col_idx, cell_data in enumerate(row_data[1:], 2):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        cell.alignment = wrap_align

        if col_idx == 2:  # "Ours" column
            cell.fill = our_fill

        if cell_data is None:
            cell.value = ""
        elif len(cell_data) == 3:
            text, font, fill = cell_data
            cell.value = text
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
        else:
            text, font = cell_data
            cell.value = text
            if font:
                cell.font = font

    ws1.row_dimensions[row_idx].height = 45

# Column widths
col_widths = [14, 22, 22, 30, 28, 32, 26, 26]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# =========================================================================
# Sheet 2: Code & Reproducibility
# =========================================================================
ws2 = wb.create_sheet("Code & Reproducibility")

headers2 = ["Model", "Code Available?", "Checkpoints?", "Can We Retrain?", "Notes"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = thin_border

repro_rows = [
    ("NLS-HAR", "No", "No", "No",
     "Not released. Author GitHub (harkash) has no HAR repo."),
    ("GOAT", "No (placeholder)", "No", "No",
     "github.com/wdkhuans/GOAT has 1 commit, README only. Empty since 2024."),
    ("LanHAR", "Yes (Apache-2.0)", "No (train from scratch)", "Yes",
     "github.com/DASHLab/LanHAR. Needs LLM API to generate activity descriptions."),
    ("CrossHAR", "Yes", "No (train from scratch)", "Yes",
     "github.com/kingdomrush2/CrossHAR. PyTorch 1.12. 19 stars."),
    ("LIMU-BERT", "Yes (MIT)", "Yes (.pt files)", "Yes (supervised only)",
     "github.com/dapowan/LIMU-BERT-Public. Has pretrained encoder + classifiers. No zero-shot capability."),
    ("IMU2CLIP", "Archived (CC-BY-NC)", "No (promised, never released)", "Impractical",
     "github.com/facebookresearch/imu2clip. Archived Jul 2025. Needs Ego4D DUA. No standard HAR benchmarks."),
]

for row_idx, (model, code, ckpt, retrain, notes) in enumerate(repro_rows, 2):
    for col_idx, val in enumerate([model, code, ckpt, retrain, notes], 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(size=10, bold=(col_idx == 1))
        cell.border = thin_border
        cell.alignment = wrap_align
    ws2.row_dimensions[row_idx].height = 36

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 24
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 60

# =========================================================================
# Sheet 3: Paper & Repo Links
# =========================================================================
ws3 = wb.create_sheet("Paper & Repo Links")

headers3 = ["Model", "Venue", "Paper URL", "Code URL", "Notes"]
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = thin_border

links = [
    ("NLS-HAR", "AAAI 2025",
     "https://arxiv.org/abs/2408.12023",
     "not released",
     "Author: Haresamudram et al."),
    ("GOAT", "IMWUT 2024",
     "https://dl.acm.org/doi/10.1145/3699736",
     "https://github.com/wdkhuans/GOAT (empty placeholder)",
     "Author: Miao & Chen. Results from different text encoders (CLIP/BERT)."),
    ("LanHAR", "IMWUT 2025",
     "https://arxiv.org/abs/2410.00003",
     "https://github.com/DASHLab/LanHAR",
     "Author: Yan et al. (Lehigh). Single-source transfer protocol."),
    ("CrossHAR", "IMWUT 2024",
     "https://dl.acm.org/doi/10.1145/3659597",
     "https://github.com/kingdomrush2/CrossHAR",
     "Author: Hong et al. (Rutgers). Our cited numbers (78.26/73.67) from HAR-DoReMi reproduction (arxiv.org/abs/2503.13542), not original paper."),
    ("LIMU-BERT", "SenSys 2021",
     "https://dl.acm.org/doi/10.1145/3485730.3485937",
     "https://github.com/dapowan/LIMU-BERT-Public",
     "Author: Xu et al. (NTU). Supervised only, no zero-shot. Has pretrained checkpoints."),
    ("IMU2CLIP", "EMNLP 2023",
     "https://arxiv.org/abs/2210.14395",
     "https://github.com/facebookresearch/imu2clip (archived)",
     "Author: Moon et al. (Meta). Ego4D only, no standard HAR benchmarks. Archived Jul 2025."),
]

for row_idx, (model, venue, paper, code, notes) in enumerate(links, 2):
    for col_idx, val in enumerate([model, venue, paper, code, notes], 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(size=10, bold=(col_idx == 1))
        cell.border = thin_border
        cell.alignment = wrap_align
    ws3.row_dimensions[row_idx].height = 40

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 46
ws3.column_dimensions['E'].width = 55

# =========================================================================
# Sheet 4: Protocol Legend
# =========================================================================
ws4 = wb.create_sheet("Protocol Legend")

legend_header = ["Term", "Meaning"]
for col, h in enumerate(legend_header, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

legend_rows = [
    ("closed-set", "Argmax over target dataset's own C labels only. No group/synonym mapping. Exact string match."),
    ("zero-shot", "Model never sees any data (labeled or unlabeled) from target dataset during training."),
    ("few-shot (~7% target)", "GOAT protocol: fine-tunes on ~7% of target dataset's labeled data after pre-training on other datasets."),
    ("4-act", "Evaluated on only 4 shared activities. LanHAR: walking, upstairs, downstairs, sitting. CrossHAR: still (sit+stand merged), walking, upstairs, downstairs."),
    ("single-src (UCI)", "Trained on 1 source dataset only (specified in parens). E.g. 'UCI' = trained on UCI-HAR."),
    ("multi-src (3ds)", "Trained on 3 source datasets combined. E.g. CrossHAR MotionSense = trained on HHAR+Shoaib+UCI."),
    ("supervised, 1% labels", "Within-dataset training using 1% of labeled data. NOT cross-dataset or zero-shot."),
    ("in training set", "This dataset is in our training data. Cannot do zero-shot comparison on it."),
    ("incompatible labels", "Baseline uses fundamentally different activity labels than us on this dataset."),
    ("pending", "Dataset is in our benchmark script. Awaiting 100-epoch model run."),
]

for row_idx, (term, meaning) in enumerate(legend_rows, 2):
    cell1 = ws4.cell(row=row_idx, column=1, value=term)
    cell1.font = Font(size=10, bold=True)
    cell1.border = thin_border
    cell2 = ws4.cell(row=row_idx, column=2, value=meaning)
    cell2.font = Font(size=10)
    cell2.border = thin_border
    cell2.alignment = wrap_align
    ws4.row_dimensions[row_idx].height = 30

ws4.column_dimensions['A'].width = 24
ws4.column_dimensions['B'].width = 80

# Save
output_path = "/home/alex/code/tsfm/baseline_comparison.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
